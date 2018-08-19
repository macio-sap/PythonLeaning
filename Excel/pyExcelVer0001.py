# -*- coding: utf-8 -*-
# [python3 读写Excel](https://blog.csdn.net/jeikerxiao/article/details/73614335)
# 说明
# 2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写。
# 2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写。
# pypi的地址：
# https://pypi.python.org/pypi/xlwt
# https://pypi.python.org/pypi/xlrd
# https://pypi.python.org/pypi/openpyxl
#
# openpyxl文档地址：
# https://openpyxl.readthedocs.io/en/latest/changes.html

import xlrd
import xlwt
# 读写2007 excel
import openpyxl


# 定义关于Excel2003 数据写入的类
def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("2003测试表")
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入Excel2003数据成功！")
    print('-+' * 20)


# 定义关于Excel2003 数据读取的类
def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t|", end="")
        print()


# 定义关于Excel2007 数据读取的类
def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入Excel2007数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    # 直接使用wb[SheetName]的方式 比 wb.get_sheet_by_name(SheetName) 更简洁
    # sheet = wb.get_sheet_by_name('2007测试表')
    sheet = wb['2007测试表']

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t|", end="")
        print()


file_2003 = 'data/2003测试文件.xls'
file_2007 = 'data/2007测试文件.xlsx'


def main():
    write03Excel(file_2003)
    read03Excel(file_2003)

    write07Excel(file_2007)
    read07Excel(file_2007)


if __name__ == '__main__':
    main()