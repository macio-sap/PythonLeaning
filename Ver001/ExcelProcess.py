#! python3
# -*- coding: utf-8 -*-
import logging.handlers
import os
import time
import xlrd
from openpyxl import workbook as wkb
from openpyxl import load_workbook
from ExcelReader.Lib import TestTkinter as tt, DirRead as dr
from pathlib import PureWindowsPath as pwt


# 读取文件信息清单
def getlist(src_dir):
    fl = dr.FileLst()
    if len(src_dir) == 0:
        base_dir = dr.get_cur_dir()
    else:
        base_dir = src_dir
    list_file = fl.getlist(base_dir)
    return list_file


# 读取文件信息清单
def crt_export_dir(log):
    export_dir = pwt(dr.get_cur_dir()).joinpath('EXPORT')
    log.info('检查目录：' + str(export_dir))
    if not os.path.exists(export_dir):
        os.mkdir(export_dir, mode=755)
        log.info('创建目录：' + str(export_dir) + '成功')
    return export_dir


# 返回读取到的XLSX文件对象
def open_xlsx(file_nam):
    wb = load_workbook(filename=file_nam, read_only=False)
    return wb


# 返回读取到的XLS文件对象
def open_xls(file_nam):
    wb = xlrd.open_workbook(file_nam)
    return wb


# 获取工作簿所有工作表名
def get_xlsx_sheet_names(wb):
    sheet_names = wb.get_sheet_names()
    return sheet_names


# 获取工作簿所有工作表名
def get_xls_sheet_names(wb):
    sheet_names = wb.sheet_names()
    return sheet_names


# 获取工作簿所有工作表名
def get_xls_sheet_by_name(wb, sheet_name):
    sh = wb.sheet_by_name(sheet_name)
    return sh


# 获取ECO信息
def get_eco_info(wb, log):
    sh = wb.sheet_by_name('工程更改申请单')
    row_id_eco = 0
    max_row = sh.nrows
    log.info('最大数据列数为:' + str(max_row))
    eco_info = []
    for row in range(1, max_row):
        if sh.cell(row, 0).value == '通知单编号':
            row_id_eco = row
        if row_id_eco == row - 1:
            eco_num = sh.cell(row, 0).value
            creator = sh.cell(row, 2).value
            cre_date = sh.cell(row, 4).value
            log.info('通知单编号:' + eco_num)
            log.info('提交人姓名:' + creator)
            log.info('提交日期:' + cre_date)
            eco_info.append(eco_num)
            eco_info.append(creator)
            eco_info.append(cre_date)
        if sh.cell(row, 0).value == '更改涉及的BOM':
            top_row = row + 1
            log.info('数据信息抬头行:' + str(top_row))
            eco_info.append(top_row)
            break
    return eco_info


# 获取ECO信息
def get_change_bom_lst(wb, eco_info, log):
    sh = wb.sheet_by_name('工程更改申请单')
    date_row = []
    max_row = sh.nrows
    lst_row = 0
    for row in range(eco_info[3], max_row):
        row_dat = sh.row_values(row)
        for index, dat in enumerate(row_dat):
            if index == 1 and len(dat) == 0:
                # log.info('最后一行的行号:' + str(row))
                lst_row = row
                break
        if lst_row > 0:
            break
        date_row.append(row_dat)
    log.info('数据条目数:' + str(len(date_row)))
    return date_row


# 获取指定Sheet 的数据内容
def get_xls_info_lst(wb, sheet_name, log):
    sh = wb.sheet_by_name(sheet_name)
    date_row = []
    if sheet_name == '更改凭证涉及的物料信息':
        min_row = 2
    elif sheet_name == '删除总成中的自制件和外购件':
        min_row = 1
    elif sheet_name == '新增总成中的自制件和外购件':
        min_row = 1
    max_row = sh.nrows
    lst_row = 0
    for row in range(min_row, max_row):
        row_dat = sh.row_values(row)
        for index, dat in enumerate(row_dat):
            if index == 1 and len(dat) == 0:
                lst_row = row
                break
        if lst_row > 0:
            break
        date_row.append(row_dat)
    log.info('数据条目数:' + str(len(date_row)))
    return date_row


# 创建写入文件对象 目标数据格式为新版本Excel
def crt_tgt_xlsx_workbook(log):
    wb = wkb.Workbook()
    wb.active
    log.info('创建目标文件成功')
    return wb


# 添加数据到文件对象 目标数据格式为新版本Excel
def apd_tgt_xlsx_workbook(ws, date_row, log):
    for col in range(1, 27):  # 写入表头
        _ = ws.cell(row=1, column=col, value=str(date_row[0][col - 1]))

    # 写入数据内容
    # row = 1
    # col = 0
    for row in range(2, len(date_row)):
        for col in range(1, 27):

            # _ = ws.cell(row=row, column=col, value=str(date_row[row][col - 1]))
            _ = ws.cell(row=row, column=col, value=date_row[row][col - 1])
            # log.info('当前行:' + str(row) + '当前列' + str(col) + '内容:' + str(date_row[row][col - 1]))


# 添加数据到文件对象 目标数据格式为新版本Excel
def apd_tgt_chg_doc_mat_info(ws, date_row, log):
    head_rowdat = date_row[0]
    max_col = len(head_rowdat)
    for col in range(1, max_col):  # 写入表头
        _ = ws.cell(row=1, column=col, value=str(date_row[0][col - 1]))

    # 写入数据内容
    for row in range(2, len(date_row)):
        for col in range(1, max_col):
            _ = ws.cell(row=row, column=col, value=date_row[row][col - 1])


# 输出工程变更申请单信息
def export_change_bom_lst(wb, tgt_dir, eco_info, log):
    date_row = get_change_bom_lst(wb, eco_info, log)
    tgt_wb = crt_tgt_xlsx_workbook(log)
    tgt_ws = tgt_wb.create_sheet('工程更改申请单')
    del tgt_wb['Sheet']
    apd_tgt_xlsx_workbook(tgt_ws, date_row, log)
    sav_tgt_xlsx_workbook(tgt_wb, tgt_dir, '工程更改申请单-' + eco_info[0] + '.xlsx', log)


# 输出更改凭证涉及的物料信息
def export_chg_doc_mat_info(wb, tgt_dir, sheet_name, eco_info, log):
    date_row = get_xls_info_lst(wb, sheet_name, log)
    tgt_wb = crt_tgt_xlsx_workbook(log)
    tgt_ws = tgt_wb.create_sheet(sheet_name)
    del tgt_wb['Sheet']
    apd_tgt_chg_doc_mat_info(tgt_ws, date_row, log)
    sav_tgt_xlsx_workbook(tgt_wb, tgt_dir, sheet_name + '-' + eco_info[0] + '.xlsx', log)


# 保存文件对象 目标数据格式为新版本Excel
def sav_tgt_xlsx_workbook(wb, tgt_dir, file_name, log):
    if len(tgt_dir) == 0:
        export_dir = crt_export_dir(log)
    else:
        export_dir = tgt_dir
    tgt_fil = pwt(export_dir).joinpath(file_name)
    log.info(str(tgt_fil))
    wb.save(filename=str(tgt_fil))
    wb.close()


# 定义日志记录相关信息
def set_log():
    log_file: str = r'tst.log'
    handler = logging.handlers.RotatingFileHandler(log_file, maxBytes=1024 * 1024, backupCount=5,
                                                   encoding='utf-8')  # 实例化handler
    fmt = '%(asctime)s - %(levelname)s - %(message)s'
    formatter = logging.Formatter(fmt)  # 实例化formatter
    handler.setFormatter(formatter)  # 为handler添加formatter

    logger = logging.getLogger('tst')  # 获取名为tst的logger
    logger.addHandler(handler)  # 为logger添加handler
    logger.setLevel(logging.DEBUG)
    return logger


def exec_file_export(src_dir, tgt_dir, log):
    cnt_cur_file = 0
    cnt_all_file = len(getlist(src_dir))
    for index, fil in enumerate(getlist(src_dir)):  # fil in getlist():
        log.info(fil)
        cnt_cur_file += 1
        # if 0 < index < 5:
        # if "500000280277" in fil :
        wb = open_xls(fil)
        print(time.asctime(time.localtime(time.time())) + ' '+ fil)
        eco_info = get_eco_info(wb, log)
        export_change_bom_lst(wb, tgt_dir, eco_info, log)
        export_chg_doc_mat_info(wb, tgt_dir, '更改凭证涉及的物料信息', eco_info, log)
        export_chg_doc_mat_info(wb, tgt_dir, '删除总成中的自制件和外购件', eco_info, log)
        export_chg_doc_mat_info(wb, tgt_dir, '新增总成中的自制件和外购件', eco_info, log)
        tt.p
        tt.pbv.set(cnt_cur_file / cnt_all_file)
    print('处理文件数：' + str(cnt_cur_file))
    return cnt_cur_file


# 主处理程序过程
def main():
    start_tm = time.asctime(time.localtime(time.time()))
    tt.ini_screen()
    end_tm = time.asctime(time.localtime(time.time()))
    print('开始时间：' + start_tm)
    print('结束时间：' + end_tm)


# 主要处理
if __name__ == '__main__':
    main()
